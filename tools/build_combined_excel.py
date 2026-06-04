# Combina las preguntas de ambos Excel de ejemplo en un unico archivo
# con el formato: Pregunta | Respuesta A..D | Respuesta correcta (letra) | Valor respuesta correcta
# Una hoja por tematica (Idioms, Phrasal Verbs, Grammar & Vocabulary) para
# poder mostrar en la web un test de cada tipo.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SOURCES = [
    "TestsExcelEjemplo/excel_piloto_test_ingles.xlsx",
    "TestsExcelEjemplo/excel_test_ingles_1-9.xlsx",
]
OUT = "banco_preguntas_combinado.xlsx"
HEADER = ["Pregunta", "Respuesta A", "Respuesta B", "Respuesta C",
          "Respuesta D", "Respuesta correcta", "Valor respuesta correcta"]

def section_of(raw):
    s = (raw or "").lower()
    if "idiom" in s:
        return "Idioms"
    if "phrasal" in s:
        return "Phrasal Verbs"
    return "Grammar & Vocabulary"   # examenes / grammar / vocab

# Recolecta y deduplica
buckets = {"Idioms": [], "Phrasal Verbs": [], "Grammar & Vocabulary": []}
seen = set()
for src in SOURCES:
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Banco completo"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        q = row[6]
        if not q or not str(q).strip():
            continue
        opts = [row[7], row[8], row[9], row[10]]
        opts = ["" if c is None else str(c).strip() for c in opts]
        letter = "" if row[11] is None else str(row[11]).strip().upper()
        value = "" if row[12] is None else str(row[12]).strip()
        sec = section_of(row[1])
        key = (sec, str(q).strip().lower(), tuple(o.lower() for o in opts))
        if key in seen:
            continue
        seen.add(key)
        buckets[sec].append([str(q).strip(), *opts, letter, value])

# Escribe el Excel
wb = openpyxl.Workbook()
wb.remove(wb.active)
hfill = PatternFill("solid", fgColor="1A365D")
hfont = Font(bold=True, color="FFFFFF")
for sec in ["Idioms", "Phrasal Verbs", "Grammar & Vocabulary"]:
    ws = wb.create_sheet(sec[:31])
    ws.append(HEADER)
    for c in ws[1]:
        c.fill = hfill; c.font = hfont; c.alignment = Alignment(vertical="center")
    for r in buckets[sec]:
        ws.append(r)
    widths = [70, 22, 22, 22, 22, 16, 26]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
wb.save(OUT)

total = sum(len(v) for v in buckets.values())
print(f"OK -> {OUT}  ({total} preguntas)")
for sec, v in buckets.items():
    print(f"  {sec}: {len(v)}")
